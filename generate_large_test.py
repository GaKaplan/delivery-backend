import openpyxl
from openpyxl.styles import Font

def create_large_excel(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja de Ruta 50"

    # Headers
    headers = ["Nro", "Dirección", "Localidad", "Bultos"]
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = text
        cell.font = Font(bold=True)

    # 50 Real addresses in Buenos Aires (CABA and GBA)
    addresses = [
        "Av. Corrientes 1200", "Florida 100", "Av. Santa Fe 2500", "Av. Cabildo 1500", "Av. Rivadavia 5000",
        "Av. Juan B. Justo 3000", "Av. San Martín 4000", "Av. Pueyrredón 100", "Av. Callao 800", "Av. Entre Ríos 1200",
        "Av. Belgrano 1500", "Av. de Mayo 600", "Av. 9 de Julio 1000", "Av. Scalabrini Ortiz 1800", "Av. Córdoba 3500",
        "Av. Paseo Colón 200", "Av. Libertador 2000", "Av. Figueroa Alcorta 3000", "Paraguay 1500", "Tucumán 800",
        "Lavalle 500", "Sarmiento 1500", "Viamonte 2000", "Av. Gaona 2500", "Luis Maria Campos 100",
        "Av. Monroe 3000", "Av. Olazábal 2500", "Av. Congreso 1500", "Av. Ricardo Balbín 2000", "Av. Triunvirato 3500",
        "Av. Federico Lacroze 2500", "Av. Alvarez Thomas 1200", "Av. de los Incas 3000", "Av. Elcano 3500", "Av. Forest 1500",
        "Av. Dorrego 1000", "Av. Corrientes 6000", "Av. Warnes 1200", "Av. Angel Gallardo 500", "Av. Estado de Israel 4200",
        "Av. Boedo 100", "Av. Chiclana 3000", "Av. Caseros 2500", "Av. Montes de Oca 1000", "Av. Patricios 500",
        "Av. Regimiento de Patricios 1000", "Av. Martin Garcia 500", "Av. Paseo Colón 1200", "Av. Brasil 1000", "Av. Garay 2000"
    ]

    for i, addr in enumerate(addresses, 2):
        ws.cell(row=i, column=1).value = i - 1
        ws.cell(row=i, column=2).value = f"{addr}, CABA"
        ws.cell(row=i, column=4).value = (i % 5) + 1

    wb.save(filename)
    print(f"Created {filename} with 50 addresses.")

if __name__ == "__main__":
    create_large_excel("test_50_addresses.xlsx")
