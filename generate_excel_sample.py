import openpyxl
from openpyxl import Workbook

def create_sample_excel(filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja de Ruta"
    
    # Fill with some dummy data to simulate the user's file structure
    ws["A7"] = "Hoja de Ruta 7024"
    ws["A9"] = "Desde 20/01/2026..."
    
    # Header Row (approx Row 15 based on user input, data starts 16)
    headers = ["Cod.Cli", "Cliente", "Comprobante", "Importe"]
    for i, h in enumerate(headers):
        ws.cell(row=15, column=i+1, value=h)
        
    # Data starting Row 16, Column B (Index 2) is "Cliente" (Address)
    data = [
        "AV GAONA 2759",
        "AV SAN MARTIN 7170",
        "CARLOS ANTONIO LOPEZ 3585",
        "CARLOS CALVO 4251",
        "CHARRUA 3172",
        "CORDOBA 4001",
        "GASCON 285"
    ]
    
    start_row = 16
    for i, address in enumerate(data):
        row_idx = start_row + i
        # Col B is 2
        ws.cell(row=row_idx, column=2, value=address)
        # Fill Col A just for display
        ws.cell(row=row_idx, column=1, value=f"Code {i}")
        
    wb.save(filename)
    print(f"Created {filename}")

if __name__ == "__main__":
    create_sample_excel("sample_delivery.xlsx")
